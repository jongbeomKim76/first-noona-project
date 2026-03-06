import openpyxl
from openpyxl.styles import Font, Alignment

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI 환경 구축 가이드"

# 헤더 설정
headers = ["카테고리", "도구/프레임워크", "설명", "설치/사용 방법", "참고"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 데이터 입력 (스터디 노트 기반)
data = [
    ["언어", "Python", "AI 개발의 주력 언어, 데이터 처리 및 LLM 연동에 강함", "pip install python", "데이터 분석 라이브러리 (pandas, numpy) 함께 사용"],
    ["언어", "TypeScript", "프론트엔드 및 백엔드에서 타입 안전한 개발", "npm install typescript", "Node.js 환경에서 주로 사용"],
    ["프레임워크", "LangChain", "LLM 체인, 에이전트, RAG 구축", "pip install langchain", "Python/JS 버전 모두 지원"],
    ["프레임워크", "Pydantic", "데이터 검증 및 모델링", "pip install pydantic", "JSON 스키마 정의에 유용"],
    ["인프라", "Docker", "컨테이너화로 환경 일관성 유지", "docker --version (설치 후)", "AI 앱 배포에 필수"],
    ["인프라", "Hugging Face Hub", "오픈소스 모델 호스팅 및 다운로드", "pip install huggingface_hub", "Transformers 라이브러리와 연동"],
    ["인프라", "Ollama", "로컬 LLM 실행 (Llama3 등)", "ollama pull llama3", "API 서버로 띄워 테스트"],
    ["모니터링", "LangSmith", "LangChain 체인 모니터링", "pip install langsmith", "토큰/비용 추적"],
    ["모니터링", "Weights & Biases", "ML/LLM 실험 기록", "pip install wandb", "대시보드 제공"],
    ["백엔드", "Spring Boot", "Java 기반 AI 백엔드", "Spring Initializr로 프로젝트 생성", "Spring AI로 LLM 연동"],
    ["프론트엔드", "React", "UI 구축, Vercel AI SDK로 스트리밍", "npx create-react-app", "SSE로 실시간 응답"],
    ["Vector DB", "Pinecone", "벡터 저장 및 검색", "pip install pinecone-client", "RAG에 사용"],
    ["Vector DB", "Milvus", "오픈소스 벡터 DB", "Docker로 실행", "대규모 데이터에 적합"],
]

# 데이터 입력
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num).value = value

# 열 너비 자동 조정
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 파일 저장
wb.save("AI_환경_구축_가이드.xlsx")
print("엑셀 파일 'AI_환경_구축_가이드.xlsx'가 생성되었습니다.")